from django.shortcuts import render, HttpResponse, HttpResponseRedirect
from django.contrib import messages
from django.utils.datastructures import MultiValueDictKeyError
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm, SetPasswordForm
from django.contrib.auth.decorators import user_passes_test
from django.forms.formsets import formset_factory
from django.db import IntegrityError
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from datetime import datetime
from pytz import timezone
from openpyxl import Workbook
import time

from loads.forms import DateFilter, UserFilter, FuelFilter, AddKeysForm
from loads.models import KeyOwner, Database, Cistern, UpDosed, FuelType, Connect
from loads.views import recalc, refr_loads, refr_keys, add_key, dev, ready
from .forms import EditKeyOwnerForm, AddSystemUserForm, EditDjangoUserForm, CisternForm, AddUpDosedForm, FuelForm, \
    BaseFuelFormSet

kiev = timezone('Europe/Kiev')


# БД отгрузок
@user_passes_test(lambda u: u.is_superuser)
def dosing(request):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    start_date = kiev.localize(datetime.utcfromtimestamp(0))
    end_date = kiev.localize(datetime.now())
    filter_name = ''
    filter_car = ''
    filter_fuel = ''
    date_filter = DateFilter(request.GET)
    user_filter = UserFilter(request.GET)
    fuel_filter = FuelFilter(request.GET)
    if date_filter.is_valid():
        if date_filter.cleaned_data['start_date']:
            start_date = kiev.localize(datetime.combine(date_filter.cleaned_data['start_date'], datetime.min.time()))
        if date_filter.cleaned_data['end_date']:
            end_date = kiev.localize(datetime.combine(date_filter.cleaned_data['end_date'], datetime.max.time()))
    if user_filter.is_valid():
        filter_name = user_filter.cleaned_data['filter_name']
        filter_car = user_filter.cleaned_data['filter_car']
    if fuel_filter.is_valid():
        filter_fuel = fuel_filter.cleaned_data['filter_fuel']
    db = Database.objects.filter(date_time__gte=start_date, date_time__lte=end_date,
                                 user__name__icontains=filter_name, user__car__icontains=filter_car,
                                 user__cistern__fuel__name__icontains=filter_fuel, delete=False)
    if request.GET.get('refr_log'):
        refr_loads(request)
    if request.GET.get('delete'):
        db.update(delete=True)
    if request.GET.get('to_xls'):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="loads_report.xls"'
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Отгружено', 'Дата', 'Цистерна', 'Тип жидкости']
        ws.append(xls_mas)
        for entry in db:
            xls_mas = [entry.user.name, entry.user.car, entry.user.keys, entry.dosed, entry.date_time]
            if entry.user.cistern:
                xls_mas.extend([entry.user.cistern.name, entry.user.cistern.fuel.name])
            else:
                xls_mas.extend(['', ''])
            ws.append(xls_mas)
        wb.save(response)
        return response
    db_paginator = Paginator(db, 25)
    db_page = request.GET.get('db_page')
    try:
        downdosed = db_paginator.page(db_page)
    except PageNotAnInteger:
        downdosed = db_paginator.page(1)
    except EmptyPage:
        downdosed = db_paginator.page(db_paginator.num_pages)
    return render(request, 'local_admin/dosing.html',
                  {'db': downdosed, 'user_filter': user_filter, 'date_filter': date_filter, 'fuel_filter': fuel_filter,
                   'cur_user': request.user.username})


# БД ключей
@user_passes_test(lambda u: u.is_superuser)
def keys(request):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    filter_name = ''
    filter_car = ''
    user_filter = UserFilter(request.GET)
    if user_filter.is_valid():
        filter_name = user_filter.cleaned_data['filter_name']
        filter_car = user_filter.cleaned_data['filter_car']
    udb = KeyOwner.objects.filter(car__icontains=filter_car, name__icontains=filter_name)
    if request.GET.get('refr_keys'):
        refr_keys(request)
    if request.GET.get('to_xls'):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="users_report.xls"'
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Комментарий']
        ws.append(xls_mas)
        for entry in udb:
            xls_mas = [entry.name, entry.car, entry.keys, entry.comment]
            ws.append(xls_mas)
        wb.save(response)
        return response
    if request.method == 'POST':
        add_keys_form = AddKeysForm(request.POST, request.FILES)
        try:
            key_doc = request.FILES['key_file'].read().decode().split()
        except MultiValueDictKeyError:
            messages.add_message(request, messages.ERROR, 'Ошибка файла')
            return HttpResponseRedirect('/admin/keys/')
        add_key(request, key_doc)
    else:
        add_keys_form = AddKeysForm()
    return render(request, 'local_admin/keys.html',
                  {'udb': udb, 'user_filter': user_filter, 'add_keys': add_keys_form,'cur_user': request.user.username})


# Редактирование БД ключей
@user_passes_test(lambda u: u.is_superuser)
def edit_key(request, key_id):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    owner = KeyOwner.objects.get(id=key_id)
    if request.method == 'POST':
        edit_form = EditKeyOwnerForm(request.POST, instance=owner)
        if edit_form.is_valid():
            edit_form.save()
            messages.add_message(request, messages.SUCCESS, 'Информация успешно отредактирована')
            return HttpResponseRedirect('/admin/keys/')
    else:
        edit_form = EditKeyOwnerForm(instance=owner)
    return render(request, 'local_admin/key_edit.html', {'edit_form': edit_form, 'cur_user': request.user.username})


# Пользователи системы
@user_passes_test(lambda u: u.is_superuser)
def users(request):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    last_name = ''
    user_filter = UserFilter(request.GET)
    if user_filter.is_valid():
        last_name = user_filter.cleaned_data['filter_name']
    if last_name:
        udb = User.objects.filter(last_name__icontains=last_name)
    else:
        udb = User.objects.all()
    disable = request.POST.get('user_name')
    if disable:
        User.objects.get(username=disable).delete()
    return render(request, 'local_admin/users.html',
                  {'udb': udb, 'cur_user': request.user.username, 'user_filter': user_filter})


# Регистрация пользователя
@user_passes_test(lambda u: u.is_superuser)
def add_local_user(request):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    if request.method == 'POST':
        general = UserCreationForm(request.POST)
        additional = AddSystemUserForm(request.POST)
        if general.is_valid() and additional.is_valid():
            e_mail = additional.cleaned_data['email']
            try:
                User.objects.get(email=e_mail)
                messages.add_message(request, messages.ERROR,
                                     'Пользователь с почтовым ящиком ' + e_mail + ' уже зарегистрирован')
            except User.DoesNotExist:
                user = general.save()
                user.first_name = additional.cleaned_data['first_name']
                user.last_name = additional.cleaned_data['last_name']
                user.email = e_mail
                user.save()
                messages.add_message(request, messages.SUCCESS,
                                     'Пользователь успешно зарегистрирован')
                return HttpResponseRedirect('/admin/users/')
    else:
        general = UserCreationForm()
        additional = AddSystemUserForm()
    return render(request, 'local_admin/user_add.html',
                  {'general': general, 'additional': additional, 'cur_user': request.user.username})


# Редактирование пользователя
@user_passes_test(lambda u: u.is_superuser)
def edit_local_user(request, user_id):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    u = User.objects.get(id=user_id)
    if request.method == 'POST':
        edit_form = EditDjangoUserForm(request.POST, instance=u)
        set_passwd = SetPasswordForm(user=u)
        if request.POST.get('ch_user'):
            if edit_form.is_valid():
                edit_form.save()
                messages.add_message(request, messages.SUCCESS, 'Информация о пользователе успешно отредактирована.')
        if request.POST.get('set_passwd'):
            set_passwd = SetPasswordForm(data=request.POST, user=u)
            if set_passwd.is_valid():
                set_passwd.save()
                messages.add_message(request, messages.SUCCESS, 'Пароль успешно изменен.')
    else:
        edit_form = EditDjangoUserForm(instance=u)
        set_passwd = SetPasswordForm(user=u)
    return render(request, 'local_admin/user_edit.html',
                  {'edit_form': edit_form, 'cur_user': request.user.username,
                   'e_user': u.username, 'set_passwd': set_passwd})


@user_passes_test(lambda u: u.is_superuser)
def add_fuel(request):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    FuelFormSet = formset_factory(FuelForm, formset=BaseFuelFormSet, can_delete=False)
    if request.method == 'POST':
        fuel_formset = FuelFormSet(request.POST)
        if fuel_formset.is_valid():
            for fuel_form in fuel_formset:
                new_fuel = FuelType(name=fuel_form.cleaned_data.get('name'))
                if fuel_form.cleaned_data.get('comment'):
                    new_fuel.comment = fuel_form.cleaned_data.get('comment')
                try:
                    new_fuel.save()
                except IntegrityError:
                    messages.add_message(request, messages.ERROR, 'Данный вид топлива уже зарегистрирован.')
        return HttpResponseRedirect('/admin/fuels/')
    else:
        fuel_formset = FuelFormSet()
    return render(request, 'local_admin/fuel_add.html', {'cur_user': request.user.username, 'fuel_formset': fuel_formset})


@user_passes_test(lambda u: u.is_superuser)
def fuels_list(request):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    fuels = FuelType.objects.all()
    if request.method == 'POST':
        edit_fuel = FuelForm(request.POST)
        if edit_fuel.is_valid():
            fuel = FuelType.objects.get(id=request.POST.get('fuel_id'))
            fuel.name = edit_fuel.cleaned_data['name']
            if edit_fuel.cleaned_data['comment']:
                fuel.comment = edit_fuel.cleaned_data['comment']
            try:
                fuel.save()
                messages.add_message(request, messages.SUCCESS, 'Информация успешно отредактирована.')
            except IntegrityError:
                messages.add_message(request, messages.ERROR, 'Данный вид топлива уже зарегистрирован.')
    else:
        edit_fuel = FuelForm()
    if request.GET.get('del'):
        fuel_del = FuelType.objects.get(id=request.GET.get('del'))
        if len(fuel_del.cistern_set.all()):
            messages.add_message(request, messages.ERROR, 'Удаление невозможно. '
                                                          'Данный тип топлива задействован в резервуаре.')
        else:
            fuel_del.delete()
    return render(request, 'local_admin/fuels.html',
                  {'cur_user': request.user.username, 'fuels': fuels, 'edit_fuel': edit_fuel})


@user_passes_test(lambda u: u.is_superuser)
def fuel_info(request, fuel_id):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    fuel = FuelType.objects.get(id=fuel_id)
    start_date = kiev.localize(datetime.utcfromtimestamp(0))
    end_date = kiev.localize(datetime.now())
    filter_name = ''
    filter_car = ''
    date_filter = DateFilter(request.GET)
    user_filter = UserFilter(request.GET)
    if date_filter.is_valid():
        if date_filter.cleaned_data['start_date']:
            start_date = kiev.localize(datetime.combine(date_filter.cleaned_data['start_date'], datetime.min.time()))
        if date_filter.cleaned_data['end_date']:
            end_date = kiev.localize(datetime.combine(date_filter.cleaned_data['end_date'], datetime.max.time()))
    if user_filter.is_valid():
        filter_name = user_filter.cleaned_data['filter_name']
        filter_car = user_filter.cleaned_data['filter_car']
    db = Database.objects.filter(date_time__gte=start_date, date_time__lte=end_date,
                                 user__name__icontains=filter_name, user__car__icontains=filter_car,
                                 user__cistern__fuel=fuel, delete=False)
    if request.GET.get('refr_log'):
        refr_loads(request)
    if request.GET.get('delete'):
        db.update(delete=True)
    if request.GET.get('to_xls'):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="loads_report.xls"'
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Отгружено', 'Дата', 'Оставшийся объем', 'Цистерна']
        ws.append(xls_mas)
        for entry in db:
            xls_mas = [entry.user.name, entry.user.car, entry.user.keys, entry.dosed, entry.date_time,
                       entry.fuel_volume, entry.user.cistern.name]
            ws.append(xls_mas)
        wb.save(response)
        return response
    db_paginator = Paginator(db, 25)
    db_page = request.GET.get('db_page')
    try:
        downdosed = db_paginator.page(db_page)
    except PageNotAnInteger:
        downdosed = db_paginator.page(1)
    except EmptyPage:
        downdosed = db_paginator.page(db_paginator.num_pages)
    return render(request, 'local_admin/fuel_info.html',
                  {'db': downdosed, 'user_filter': user_filter, 'date_filter': date_filter,
                   'cur_user': request.user.username})


# Добавление резервуара
@user_passes_test(lambda u: u.is_superuser)
def cistern_add(request):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    if not KeyOwner.objects.exists():
        messages.add_message(request, messages.ERROR, 'В системе нет зарегистрированных ключей.'
                                                      'Считайте перечень ключей для продолжения работы')
        return HttpResponseRedirect('/admin/keys/')
    fuel_types = FuelType.objects.all()
    if len(fuel_types) == 0:
        messages.add_message(request, messages.WARNING, 'Не заданы типы топлива, добавление резервуара невозможно')
        return HttpResponseRedirect('/admin/fuels/add-fuel/')
    free_keys = KeyOwner.objects.filter(cistern=None)
    if len(free_keys) == 0:
        messages.add_message(request, messages.WARNING, 'Нет незадействованных ключей, добавление резервуара невозможно')
        return HttpResponseRedirect('/admin/keys/')
    if request.method == 'POST':
        add_cist = CisternForm(request.POST)
        if add_cist.is_valid():
            try:
                fuel = FuelType.objects.get(name=request.POST['select_fuel'])
                cist = add_cist.save(commit=False)
                cist.fuel = fuel
                cist.save()
                keys_list = request.POST.getlist('keys')
                for k in keys_list:
                    cur_key = KeyOwner.objects.get(keys=k)
                    cur_key.cistern = cist
                    cur_key.save()
                messages.add_message(request, messages.SUCCESS, 'Резервуар успешно добавлен')
                return HttpResponseRedirect('/admin/cisterns/')
            except ValueError:
                messages.add_message(request, messages.ERROR, 'Начальное значение не может превышать максимальное')
                return HttpResponseRedirect('/admin/cisterns/')
    else:
        add_cist = CisternForm()
    return render(request, 'local_admin/cistern_add.html',
                  {"add_cist": add_cist, 'keys': free_keys, 'fuel_types': fuel_types, 'cur_user': request.user.username})


# Перечень резервуаров
@user_passes_test(lambda u: u.is_superuser)
def cistern_list(request):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    cists = Cistern.objects.all()
    if len(cists) == 0:
        if not KeyOwner.objects.exists():
            messages.add_message(request, messages.WARNING, 'Считайте перечень ключей с устройства и '
                                                            'задайте параметры резервуара для продолжения работы')
            return HttpResponseRedirect('/admin/keys/')
        messages.add_message(request, messages.WARNING,
                             'Нет зарегистрированных резервуаров. Добавьте резервуар для продолжения работы.')
        if not FuelType.objects.exists():
            messages.add_message(request, messages.WARNING, 'Задайте типы топлива для продолжения работы.')
            return HttpResponseRedirect('/admin/fuels/add-fuel/')
        return HttpResponseRedirect('/admin/cisterns/add-cistern/')
    percents = []
    cur_vols = []
    fuels = FuelType.objects.all()
    if request.GET.get('select_fuel'):
        cists = Cistern.objects.filter(fuel__name=request.GET.get('select_fuel'))
    for c in cists:
        try:
            db = Database.objects.filter(user__cistern=c).latest('date_time')
            cur_volume = db.cistern_volume
            cur_vols.append(cur_volume)
        except Database.DoesNotExist:
            cur_volume = c.start_volume
            cur_vols.append(c.start_volume)
        perc = int(cur_volume / c.max_volume * 100 + 5 - (cur_volume / c.max_volume * 100 + 5) % 10)
        percents.append(perc)
    return render(request, 'local_admin/cisterns.html',
                  {'cists': zip(cists, cur_vols, percents), 'cur_user': request.user.username, 'fuels': fuels})


@user_passes_test(lambda u: u.is_superuser)
def cistern_edit(request, cist_id):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    c = Cistern.objects.get(id=cist_id)
    free_keys = KeyOwner.objects.filter(cistern=None)
    sel_keys = KeyOwner.objects.filter(cistern=c)
    fuel_types = FuelType.objects.exclude(id=c.fuel.id)
    if len(fuel_types) == 0:
        messages.add_message(request, messages.WARNING, 'Изменение типа топлива невозможно.')
        fuel_types = None
    if request.method == 'POST':
        edit_cist = CisternForm(request.POST, instance=c)
        try:
            if edit_cist.is_valid():
                cist = edit_cist.save(commit=False)
                fuel_recalc = [c.fuel]
                if request.POST['select_fuel'] != c.fuel.name:
                    cist.fuel = FuelType.objects.get(name=request.POST['select_fuel'])
                    fuel_recalc.append(c.fuel)
                cist.save()
                keys_list = request.POST.getlist('keys')
                KeyOwner.objects.filter(keys__in=keys_list).update(cistern=c)
                KeyOwner.objects.exclude(keys__in=keys_list).update(cistern=None)
                if Database.objects.filter(user__cistern__fuel__in=fuel_recalc).count() > 0:
                    recalc(request, c, fuel_recalc)
                messages.add_message(request, messages.SUCCESS, 'Значения успешно изменены')
                return HttpResponseRedirect('/admin/cisterns/')
        except ValueError:
            messages.add_message(request, messages.ERROR, 'Начальное значение не может превышать максимальное')
            return HttpResponseRedirect('/admin/cisterns/')
    else:
        edit_cist = CisternForm(instance=c)
    return render(request, 'local_admin/cistern_edit.html',
                  {"edit_cist": edit_cist, 'can_select': free_keys, 'selected': sel_keys, 'cur_fuel': c.fuel.name,
                   'fuel_types': fuel_types, 'cur_user': request.user.username})


@user_passes_test(lambda u: u.is_superuser)
def cistern_info(request, cist_id):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    nav = "downdosed"
    cist = Cistern.objects.get(id=cist_id)
    start_date = kiev.localize(datetime.utcfromtimestamp(0))
    end_date = kiev.localize(datetime.now())
    filter_name = ''
    filter_car = ''
    date_filter = DateFilter(request.GET)
    user_filter = UserFilter(request.GET)
    if date_filter.is_valid():
        if date_filter.cleaned_data['start_date']:
            start_date = kiev.localize(datetime.combine(date_filter.cleaned_data['start_date'], datetime.min.time()))
        if date_filter.cleaned_data['end_date']:
            end_date = kiev.localize(datetime.combine(date_filter.cleaned_data['end_date'], datetime.max.time()))
    if user_filter.is_valid():
        filter_name = user_filter.cleaned_data['filter_name']
        filter_car = user_filter.cleaned_data['filter_car']
    if request.GET.get('nav'):
        nav = request.GET.get('nav')
    downdosed = Database.objects.filter(date_time__gte=start_date, date_time__lte=end_date, delete=False,
                                        user__name__icontains=filter_name, user__car__icontains=filter_car,
                                        user__cistern=cist)
    updosed = cist.updosed_set.filter(date_time__gte=start_date, date_time__lte=end_date)
    recovery = Database.objects.filter(date_time__gte=start_date, date_time__lte=end_date, delete=True,
                                       user__name__icontains=filter_name, user__car__icontains=filter_car,
                                       user__cistern=cist)
    if request.method == 'POST':
        add_updosed = AddUpDosedForm(request.POST)
        if add_updosed.is_valid():
            load = Database.objects.get(id=request.POST.get('load_id'))
            if add_updosed.cleaned_data['volume'] + load.cistern_volume > load.cistern.max_volume:
                messages.add_message(request, messages.WARNING, 'Текущий объем не может быть больше максимального')
                return HttpResponseRedirect('/admin/cisterns/')
            add = UpDosed(user=request.user, cistern=load.user.cistern, date_time=load.date_time,
                          volume=add_updosed.cleaned_data['volume'])
            if add_updosed.cleaned_data['comment']:
                add.comment = add_updosed.cleaned_data['comment']
            add.save()
            load.add += add_updosed.cleaned_data['volume']
            load.cistern_volume += add_updosed.cleaned_data['volume']
            load.fuel_volume += add_updosed.cleaned_data['volume']
            load.save()
            next_cist_dosings = Database.objects.filter(user__cistern=load.user.cistern,
                                                        date_time__gt=load.date_time).order_by('date_time')
            if len(next_cist_dosings):
                previous_cist_volume = load.cistern_volume
                for db in next_cist_dosings:
                    db.cistern_volume = previous_cist_volume - db.dosed + db.add
                    previous_cist_volume = db.cistern_volume
                    db.save()
            next_dosings = Database.objects.filter(date_time__gt=load.date_time,
                                                   user__cistern__fuel=load.user.cistern.fuel).order_by('date_time')
            if len(next_dosings):
                previous_fuel_volume = load.fuel_volume
                for db in next_dosings:
                    db.fuel_volume = previous_fuel_volume - db.dosed + db.add
                    previous_fuel_volume = db.fuel_volume
                    db.save()
            messages.add_message(request, messages.SUCCESS, 'Загрузка успешно добавлена')
    else:
        add_updosed = AddUpDosedForm()
    if request.GET.get('delete'):
        downdosed.update(delete=True)
    if request.GET.get('recover'):
        recovery.update(delete=False)
    if request.GET.get('to_xls') and nav == "downdosed":
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = 'downdosed_' + cist_id + '.xls"'
        response['Content-Disposition'] = 'attachment; filename="' + filename
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Отгружено', 'Дата', 'Объем в цистерне', 'Добавочный объем']
        ws.append(xls_mas)
        for entry in downdosed:
            xls_mas = [entry.user.name, entry.user.car, entry.user.keys,
                       entry.dosed, entry.date_time, entry.cistern_volume, entry.add]
            ws.append(xls_mas)
        wb.save(response)
        return response
    if request.GET.get('to_xls') and nav == "updosed":
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = 'updosed_' + cist_id + '.xls"'
        response['Content-Disposition'] = 'attachment; filename="' + filename
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Фамилия', 'Объем', 'Дата', 'Комментарий']
        ws.append(xls_mas)
        for entry in updosed:
            xls_mas = [entry.user.first_name, entry.user.last_name, entry.volume, entry.date_time, entry.comment]
            ws.append(xls_mas)
        wb.save(response)
        return response
    db_paginator = Paginator(downdosed, 25)
    db_page = request.GET.get('db_page')
    try:
        db = db_paginator.page(db_page)
    except PageNotAnInteger:
        db = db_paginator.page(1)
    except EmptyPage:
        db = db_paginator.page(db_paginator.num_pages)

    ud_paginator = Paginator(updosed, 25)
    ud_page = request.GET.get('ud_page')
    try:
        ud = ud_paginator.page(ud_page)
    except PageNotAnInteger:
        ud = ud_paginator.page(1)
    except EmptyPage:
        ud = ud_paginator.page(ud_paginator.num_pages)

    rec_paginator = Paginator(recovery, 25)
    rec_page = request.GET.get('rec_page')
    try:
        rec = rec_paginator.page(rec_page)
    except PageNotAnInteger:
        rec = rec_paginator.page(1)
    except EmptyPage:
        rec = rec_paginator.page(rec_paginator.num_pages)
    return render(request, 'local_admin/cistern_info.html',
                  {'cur_user': request.user.username, 'nav': nav, 'date_filter': date_filter,
                   'user_filter': user_filter, 'downdosed': db, 'updosed': ud, 'recovery': rec,
                   'add_updosed': add_updosed})


@user_passes_test(lambda u: u.is_superuser)
def hide_admin(request):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    if request.POST.get('delete_db'):
        Connect.objects.all().delete()
        UpDosed.objects.all().delete()
        Cistern.objects.all().delete()
        Database.objects.all().delete()
        KeyOwner.objects.all().delete()
        FuelType.objects.all().delete()
        messages.add_message(request, messages.SUCCESS, 'Все данные успешно удалены.')
        return HttpResponseRedirect('/admin/hide-admin/')
    if request.POST.get('clean_log'):
        try:
            conn = Connect.objects.first()
        except Connect.DoesNotExist:
            messages.add_message(request, messages.ERROR, 'Задайте параметры подключения.')
            return HttpResponseRedirect('/admin/settings/')
        dev.port = conn.port
        dev.baudrate = conn.speed
        if not dev.isOpen():
            try:
                dev.open()
            except OSError:
                messages.add_message(request, messages.ERROR,
                                     'Невозможно установить соединение c ' + dev.port +
                                     '. Проверьте подключение устройства к компьютеру')
                return HttpResponseRedirect('/admin/')
            repeat = 0
            while repeat < 2:
                for log_on in range(len('clion\r')):
                    dev.write('clion\r'[log_on].encode())
                resp = dev.readline()
                if 'DOSA-10W' in resp.decode():
                    for log_on in range(len('clean log -f\r')):
                        dev.write('clean log -f\r'[log_on].encode())
                    resp = dev.readline()
                    if 'DOSA-10W' in resp.decode():
                        for log_on in range(len('clean log -f\r')):
                            dev.write('clean log -f\r'[log_on].encode())
                    for log_off in range(len('clioff\r')):
                        dev.write('clioff\r'[log_off].encode())
                    dev.readline()
                    dev.close()
                    repeat = 2
                else:
                    time.sleep(2)
                    repeat += 1
                    if repeat == 2:
                        dev.close()
                        messages.add_message(request, messages.ERROR,
                                             'Обмен данными c ' + dev.port +
                                             ' невозможен. Проверьте соединение устройства с RS и '
                                             'выполните подключение')
                        return HttpResponseRedirect('/admin/')
        return HttpResponseRedirect('/admin/')
    return render(request, 'local_admin/hide_admin.html')