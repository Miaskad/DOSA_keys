from django.shortcuts import render, HttpResponse, HttpResponseRedirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from openpyxl import Workbook
from datetime import datetime
from pytz import timezone

from loads.forms import DateFilter, UserFilter, FuelFilter
from loads.models import KeyOwner, Database, Cistern, FuelType
from loads.views import refr_loads, refr_keys, ready

kiev = timezone('Europe/Kiev')


# БД отгрузок
@login_required
def dosing(request):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    start_date = kiev.localize(datetime.utcfromtimestamp(0))
    end_date = kiev.localize(datetime.now())
    filter_name = ''
    filter_car = ''
    filter_fuel = ''
    date_filter = DateFilter(request.GET)
    user_filter = UserFilter(request.GET)
    fuel_filter = FuelFilter(request.GET)
    if date_filter.is_valid():
        if date_filter.cleaned_data['start_date']:
            start_date = kiev.localize(datetime.combine(date_filter.cleaned_data['start_date'], datetime.min.time()))
        if date_filter.cleaned_data['end_date']:
            end_date = kiev.localize(datetime.combine(date_filter.cleaned_data['end_date'], datetime.max.time()))
    if user_filter.is_valid():
        filter_name = user_filter.cleaned_data['filter_name']
        filter_car = user_filter.cleaned_data['filter_car']
    if fuel_filter.is_valid():
        filter_fuel = fuel_filter.cleaned_data['filter_fuel']
    db = Database.objects.filter(date_time__gte=start_date, date_time__lte=end_date,
                                 user__name__icontains=filter_name, user__car__icontains=filter_car,
                                 user__cistern__fuel__name__icontains=filter_fuel, delete=False)
    if request.GET.get('refr_log'):
        refr_loads(request)
    if request.GET.get('to_xls'):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="loads_report.xls"'
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Отгружено', 'Дата', 'Объем в цистерне',
                   'Цистерна', 'Тип жидкости']
        ws.append(xls_mas)
        for entry in db:
            xls_mas = [entry.user.name, entry.user.car, entry.user.keys,
                       entry.dosed, entry.date_time, entry.cistern_volume,
                       entry.cistern.name, entry.user.cistern.fuel.name]
            ws.append(xls_mas)
        wb.save(response)
        return response
    db_paginator = Paginator(db, 25)
    db_page = request.GET.get('db_page')
    try:
        downdosed = db_paginator.page(db_page)
    except PageNotAnInteger:
        downdosed = db_paginator.page(1)
    except EmptyPage:
        downdosed = db_paginator.page(db_paginator.num_pages)
    return render(request, 'users/user_dosing.html',
                  {'db': downdosed, 'user_filter': user_filter, 'date_filter': date_filter,
                   'fuel_filter': fuel_filter, 'cur_user': request.user.username})


# БД ключей
@login_required
def keys(request):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    filter_name = ''
    filter_car = ''
    user_filter = UserFilter(request.GET)
    if user_filter.is_valid():
        filter_name = user_filter.cleaned_data['filter_name']
        filter_car = user_filter.cleaned_data['filter_car']
    if not filter_name and not filter_car:
        udb = KeyOwner.objects.all()
    else:
        udb = KeyOwner.objects.filter(car__icontains=filter_car, name__icontains=filter_name)
    if request.GET.get('refr_keys'):
        refr_keys(request)
    if request.GET.get('to_xls'):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="users_report.xls"'
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Комментарий']
        ws.append(xls_mas)
        for entry in udb:
            xls_mas = [entry.name, entry.car, entry.keys, entry.comment]
            ws.append(xls_mas)
        wb.save(response)
        return response
    return render(request, 'users/user_keys.html',
                  {'udb': udb, 'user_filter': user_filter, 'cur_user': request.user.username})


# Перечень резервуаров
@login_required
def cistern_list(request):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    cists = Cistern.objects.all()
    if len(cists) == 0:
        messages.add_message(request, messages.ERROR,
                             'Нет зарегистрированных резервуаров. Обратитесь к администратору.')
        return HttpResponseRedirect('/user/')
    percents = []
    cur_vols = []
    fuels = FuelType.objects.all()
    if request.GET.get('select_fuel'):
        cists = Cistern.objects.filter(fuel__name=request.GET.get('select_fuel'))
    for c in cists:
        try:
            db = Database.objects.filter(user__cistern=c).latest('date_time')
            cur_volume = db.cistern_volume
            cur_vols.append(cur_volume)
        except Database.DoesNotExist:
            cur_volume = c.start_volume
            cur_vols.append(c.start_volume)
        perc = int(cur_volume / c.max_volume * 100 + 5 - (cur_volume / c.max_volume * 100 + 5) % 10)
        percents.append(perc)
    return render(request, 'users/user_cisterns.html',
                  {'cists': zip(cists, cur_vols, percents), 'cur_user': request.user.username, 'fuels': fuels})


@login_required
def cistern_info(request, cist_id):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    nav = "downdosed"
    cist = Cistern.objects.get(id=cist_id)
    start_date = kiev.localize(datetime.utcfromtimestamp(0))
    end_date = kiev.localize(datetime.now())
    filter_name = ''
    filter_car = ''
    date_filter = DateFilter(request.GET)
    user_filter = UserFilter(request.GET)
    if date_filter.is_valid():
        if date_filter.cleaned_data['start_date']:
            start_date = kiev.localize(datetime.combine(date_filter.cleaned_data['start_date'], datetime.min.time()))
        if date_filter.cleaned_data['end_date']:
            end_date = kiev.localize(datetime.combine(date_filter.cleaned_data['end_date'], datetime.max.time()))
    if user_filter.is_valid():
        filter_name = user_filter.cleaned_data['filter_name']
        filter_car = user_filter.cleaned_data['filter_car']
    if request.GET.get('nav'):
        nav = request.GET.get('nav')
    downdosed = Database.objects.filter(date_time__gte=start_date, date_time__lte=end_date, delete=False,
                                        user__name__icontains=filter_name, user__car__icontains=filter_car,
                                        user__cistern=cist)
    updosed = cist.updosed_set.filter(date_time__gte=start_date, date_time__lte=end_date)
    if request.GET.get('to_xls') and nav == "downdosed":
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = 'downdosed_' + cist_id + '.xls"'
        response['Content-Disposition'] = 'attachment; filename="' + filename
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Отгружено', 'Дата', 'Объем в цистерне', 'Добавочный объем']
        ws.append(xls_mas)
        for entry in downdosed:
            xls_mas = [entry.user.name, entry.user.car, entry.user.keys,
                       entry.dosed, entry.date_time, entry.cistern_volume, entry.add]
            ws.append(xls_mas)
        wb.save(response)
        return response
    if request.GET.get('to_xls') and nav == "updosed":
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = 'updosed_' + cist_id + '.xls"'
        response['Content-Disposition'] = 'attachment; filename="' + filename
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Фамилия', 'Объем', 'Дата', 'Комментарий']
        ws.append(xls_mas)
        for entry in updosed:
            xls_mas = [entry.user.first_name, entry.user.last_name, entry.volume, entry.date_time, entry.comment]
            ws.append(xls_mas)
        wb.save(response)
        return response
    db_paginator = Paginator(downdosed, 25)
    db_page = request.GET.get('db_page')
    try:
        db = db_paginator.page(db_page)
    except PageNotAnInteger:
        db = db_paginator.page(1)
    except EmptyPage:
        db = db_paginator.page(db_paginator.num_pages)

    ud_paginator = Paginator(updosed, 25)
    ud_page = request.GET.get('ud_page')
    try:
        ud = ud_paginator.page(ud_page)
    except PageNotAnInteger:
        ud = ud_paginator.page(1)
    except EmptyPage:
        ud = ud_paginator.page(ud_paginator.num_pages)
    return render(request, 'users/user_cistern_info.html',
                  {'cur_user': request.user.username, 'nav': nav, 'date_filter': date_filter,
                   'user_filter': user_filter, 'downdosed': db, 'updosed': ud})


@login_required
def fuel_info(request, fuel_id):
    if not ready(request):
        return HttpResponseRedirect('/licence/')
    fuel = FuelType.objects.get(id=fuel_id)
    start_date = kiev.localize(datetime.utcfromtimestamp(0))
    end_date = kiev.localize(datetime.now())
    filter_name = ''
    filter_car = ''
    date_filter = DateFilter(request.GET)
    user_filter = UserFilter(request.GET)
    if date_filter.is_valid():
        if date_filter.cleaned_data['start_date']:
            start_date = kiev.localize(datetime.combine(date_filter.cleaned_data['start_date'], datetime.min.time()))
        if date_filter.cleaned_data['end_date']:
            end_date = kiev.localize(datetime.combine(date_filter.cleaned_data['end_date'], datetime.max.time()))
    if user_filter.is_valid():
        filter_name = user_filter.cleaned_data['filter_name']
        filter_car = user_filter.cleaned_data['filter_car']
    db = Database.objects.filter(date_time__gte=start_date, date_time__lte=end_date,
                                 user__name__icontains=filter_name, user__car__icontains=filter_car,
                                 user__cistern__fuel=fuel, delete=False)
    if request.GET.get('to_xls'):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="loads_report.xls"'
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Отгружено', 'Дата', 'Оставшийся объем', 'Цистерна']
        ws.append(xls_mas)
        for entry in db:
            xls_mas = [entry.user.name, entry.user.car, entry.user.keys, entry.dosed, entry.date_time,
                       entry.fuel_volume, entry.user.cistern.name]
            ws.append(xls_mas)
        wb.save(response)
        return response
    db_paginator = Paginator(db, 25)
    db_page = request.GET.get('db_page')
    try:
        downdosed = db_paginator.page(db_page)
    except PageNotAnInteger:
        downdosed = db_paginator.page(1)
    except EmptyPage:
        downdosed = db_paginator.page(db_paginator.num_pages)
    return render(request, 'users/user_fuel_info.html',
                  {'db': downdosed, 'user_filter': user_filter, 'date_filter': date_filter,
                   'cur_user': request.user.username})
